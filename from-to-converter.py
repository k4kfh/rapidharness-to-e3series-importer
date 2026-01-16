from dataclasses import dataclass
import openpyxl
import os
import re
import click
import csv
from pathlib import Path

@dataclass(frozen=True)
class E3WireComponent:
    """Represents a valid individual wire type in the E3 database. For example, 14AWG Brown TXL."""
    wire_group : str
    """Wire Group, for example `TXL-BROWN`."""
    wire_type : str
    """Wire Type, for example `12-AWG-BRN`."""
    cross_section_awg : int
    """Wire Cross-Section in AWG. For example, `12`."""
    color : str
    """Wire Color. No rigid formatting enforced here yet. E3 seems to like 3-letter color codes like `PNK`, `RED`, `BRN`..."""

@dataclass(frozen=True)
class RapidHarnessEndpoint:
    """Represents a specific pin on a specific connector or device in RapidHarness."""
    raw_string : str
    """The raw reference designator, unmodified."""

    def get_device_des(self) -> str:
        """Return only the device/connector portion of `raw_string` (strip away the pin designation, if there is one)."""
        return self.raw_string.split(".")[0]
    
    def get_pin_des(self) -> str:
        """Return only the device/connector portion of `raw_string` (strip away the leading device designation).
        If this is a device like a Ring Terminal that has no pins, will return None."""
        split_raw_str = self.raw_string.split(".")
        if len(split_raw_str) == 1:
            # This is a "pinless" device. Return None.
            return None
        else:
            # This device must have pins. So we'll return the second item in the "split" array.
            # If there are any exceptions, let the calling code handle it.
            return split_raw_str[1]

@dataclass
class E3FromToListRow:
    # FROM RefDes
    from_device_name : str = None
    from_device_pn : str = None
    from_pin : str = None
    # TO RefDes
    to_device_name : str = None
    to_device_pn : str = None
    to_pin : str = None
    # Conductor/Signal Info
    wire : E3WireComponent = None
    signal_name : str = None
    wire_index : str = None

# WIRE LOOKUP TABLE LOADING
# Lookup tables are now loaded from CSV files at runtime

def load_wire_lookup_table(csv_path: Path) -> dict:
    """Load wire lookup table from CSV file.
    
    CSV Format:
    RapidHarness_Name,Wire_Group,E3_Wire_Type,AWG_Gauge,Color
    
    Example:
    Generic 14AWG TXL Red,TXL,14-AWG-RED,14,RED
    Generic 20AWG TXL Black,TXL,20-AWG-BLK,20,BLACK
    """
    wire_lut = {}
    with open(csv_path, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            wire_lut[row['RapidHarness_Name']] = E3WireComponent(
                wire_group=row['Wire_Group'],
                wire_type=row['E3_Wire_Type'],
                cross_section_awg=int(row['AWG_Gauge']),
                color=row['Color']
            )
    return wire_lut

def load_device_lookup_table(csv_path: Path) -> dict:
    """Load device/connector lookup table from CSV file.
    
    CSV Format:
    RapidHarness_PartNumber,E3_Device_Name
    
    Example:
    CONNECTOR-001,DT06-3S-E008
    TERMINAL-002,RingTerm_Example
    """
    device_lut = {}
    with open(csv_path, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            device_lut[row['RapidHarness_PartNumber']] = row['E3_Device_Name']
    return device_lut

def convert_rh_partnumber_to_e3(row : E3FromToListRow, device_lut: dict):
    """Given a row of the E3 From/To list, evaluate it against a set of mappings.
    If it matches any of the mappings, then 'convert' it to the mapped E3 part number.
    This helps circumvent any subtle differences in naming between E3 and RapidHarness (e.g. "DT04-3S" vs "DT043S"...that kind of thing)
    
    If no match is found in the conversion process, it just returns the original unmodified part number (or lack thereof)."""
    
    # First do the FROM device
    processed_from_device_pn = row.from_device_pn
    if row.from_device_pn in device_lut.keys():
        processed_from_device_pn = device_lut[row.from_device_pn]
    # Now check if it's most likely a splice.
    # Splices have no P/N in RapidHarness, but for E3 they should have the SPLICE part number (otherwise E3 will make them a connector and this is a nuisance to fix)
    # The pattern r'S\d+$' breaks down as:
    # S     - Matches the literal character 'S'.
    # \d+   - Matches one or more digits (0-9).
    # $     - Anchors the match to the very end of the string.
    elif re.search(r'S\d+$', row.from_device_name):
        # It probably is a splice.
        processed_from_device_pn = "SPLICE"
    
    # Next do the TO device
    processed_to_device_pn = row.to_device_pn
    if row.to_device_pn in device_lut.keys():
        processed_to_device_pn = device_lut[row.to_device_pn]
    # Same as above, check if it's a splice
    elif re.search(r'S\d+$', row.to_device_name):
        # It probably is a splice.
        processed_to_device_pn = "SPLICE"
    
    return (processed_from_device_pn, processed_to_device_pn)

def main():
    pass

@click.command()
@click.option(
    '--input', '-i',
    'input_file',
    required=True,
    type=click.Path(exists=True, dir_okay=False, readable=True, path_type=Path),
    help='Path to the RapidHarness Excel export file'
)
@click.option(
    '--output', '-o',
    'output_file',
    required=True,
    type=click.Path(dir_okay=False, writable=True, path_type=Path),
    help='Path for the output E3.series From-To List Excel file'
)
@click.option(
    '--wire-map', '-w',
    'wire_map_file',
    required=True,
    type=click.Path(exists=True, dir_okay=False, readable=True, path_type=Path),
    help='Path to the wire lookup table CSV file'
)
@click.option(
    '--device-map', '-d',
    'device_map_file',
    required=True,
    type=click.Path(exists=True, dir_okay=False, readable=True, path_type=Path),
    help='Path to the device/connector lookup table CSV file'
)
@click.option(
    '--verbose', '-v',
    is_flag=True,
    help='Enable verbose output'
)
def cli_main(input_file, output_file, wire_map_file, device_map_file, verbose):
    """Convert RapidHarness wire harness exports to E3.series From-To List format.
    
    This tool reads connection data from a RapidHarness Excel export and converts it
    to the format required by Zuken E3.series CAD software for import.
    """
    
    if verbose:
        click.echo(f"Input file: {input_file}")
        click.echo(f"Output file: {output_file}")
        click.echo(f"Wire map: {wire_map_file}")
        click.echo(f"Device map: {device_map_file}")
        click.echo("Loading lookup tables...")
    
    # Load lookup tables from CSV files
    try:
        wire_lut = load_wire_lookup_table(wire_map_file)
        if verbose:
            click.echo(f"Loaded {len(wire_lut)} wire mappings")
    except Exception as e:
        click.echo(f"Error loading wire lookup table: {e}", err=True)
        raise click.Abort()
    
    try:
        device_lut = load_device_lookup_table(device_map_file)
        if verbose:
            click.echo(f"Loaded {len(device_lut)} device mappings")
    except Exception as e:
        click.echo(f"Error loading device lookup table: {e}", err=True)
        raise click.Abort()
    
    if verbose:
        click.echo("Starting conversion...")
    
    rh_excel_path = str(input_file)
    e3_fromto_excel_path = str(output_file)

    # PHASE 1: PARSE RAPIDHARNESS DATA INTO E3 FROM/TO LIST
    rh_excel_wkbk = openpyxl.load_workbook(rh_excel_path)
    # We will store a bunch of E3FromToRows in this list
    e3_fromto = list()
    for row in range(11, rh_excel_wkbk["Connections"].max_row):
        # Create empty Row object that we'll load up with data
        e3_fromto_row = E3FromToListRow()

        # Column A: Index (skip)
            
        # Column B: FROM Designation, in Connector.Pin format
        from_endpoint = RapidHarnessEndpoint(rh_excel_wkbk["Connections"].cell(row=row, column=2).value)
        e3_fromto_row.from_device_name = from_endpoint.get_device_des()
        e3_fromto_row.from_pin = from_endpoint.get_pin_des()

        # Column C: TO Designation, in Connector.Pin format
        to_endpoint = RapidHarnessEndpoint(rh_excel_wkbk["Connections"].cell(row=row, column=3).value)
        e3_fromto_row.to_device_name = to_endpoint.get_device_des()
        e3_fromto_row.to_pin = to_endpoint.get_pin_des()

        # Column D: Conductor
        # This is the weird one in W19.Black format. We only want the index (e.g. 19 in the previous example).
        # This regex does that
        e3_fromto_row.wire_index = int(
            re.search(r'\d+',
                      rh_excel_wkbk["Connections"].cell(row=row, column=4).value
            ).group()
        )

        # Column E: Wire Part Number
        # We'll get all the info we need about the wire by mapping this part number to an E3 wire type with a lookup table
        # If this is a Cable part number (not a discrete conductor) this will throw an exception since it won't be in the LUT.
        rh_wire_sku = rh_excel_wkbk["Connections"].cell(row=row, column=5).value
        try:
            e3_fromto_row.wire = wire_lut[rh_wire_sku]
        except Exception as e:
            print(f"Unable to find wire '{rh_wire_sku}' in E3 database. Leaving cell blank in from/to list.")
        
        # Column K: FROM Connector Part Number
        e3_fromto_row.from_device_pn = rh_excel_wkbk["Connections"].cell(row=row, column=11).value

        # Column M: TO Connector Part Number
        e3_fromto_row.to_device_pn = rh_excel_wkbk["Connections"].cell(row=row, column=13).value

        # Column O: Signal Name (from Notes column in RapidHarness)
        e3_fromto_row.signal_name = rh_excel_wkbk["Connections"].cell(row=row, column=15).value

        # Store the row
        e3_fromto.append(e3_fromto_row)

    # PHASE 2 : Output E3FromToList into first sheet in workbook
    output_wb = openpyxl.Workbook()
    output_wb.active.title = "From-To List"

    # TODO: Write headers
    headers = ["From Assignment", "From Location", "From Device Name", "From Device Part #", "From Pin", "From Pin Part #", "To Assignment", "To Location", "To Device Name", "To Device Part #", "To Pin", "To Pin Part #", "Wire/Conductor Number", "Signal", "Wire Type", "Wire Color", "Wire Gauge"]
    for index,header_str in enumerate(headers,start=1):
        output_wb.active.cell(row=1, column=index, value=header_str)
    
    # WRITE ACTUAL DATA
    # The start=2 leaves the first column for headers
    for index,row in enumerate(e3_fromto, start=2):
        # Column C: From Device Name
        output_wb.active.cell(row=index, column=3, value=row.from_device_name)
        # Column D: From Device Part Number
        output_wb.active.cell(row=index, column=4, value=convert_rh_partnumber_to_e3(row, device_lut)[0])
        # Column E: From Device Pin
        output_wb.active.cell(row=index, column=5, value=row.from_pin)
        # Column I: To Device Name
        output_wb.active.cell(row=index, column=9, value=row.to_device_name)
        # Column J: To Device Part Number
        output_wb.active.cell(row=index, column=10, value=convert_rh_partnumber_to_e3(row, device_lut)[1])
        # Column K: To Device Pin
        output_wb.active.cell(row=index, column=11, value=row.to_pin)

        # WIRE INFO
        # Column M: Wire Number
        # We take the wire number from RapidHarness, and strip the W off the front of it (since E3 does not normally include W's as prefixes for wires themselves)
        output_wb.active.cell(row=index, column=13, value=row.wire_index)
        
        # Column N: Signal Name
        output_wb.active.cell(row=index, column=14, value=row.signal_name)
        
        # Conductor Data
        # May or may not be present...if it's part of a cable, this will be None since we need to handle these fully manually
        if isinstance(row.wire, E3WireComponent):
            # Column O: Wire Group (e.g. TXL, GXL, etc)
            output_wb.active.cell(row=index, column=15, value=row.wire.wire_group)
            # Column P: Wire Color
            output_wb.active.cell(row=index, column=16, value=row.wire.color)
            # Column Q: Wire Gauge in 18-AWG-ORG format
            output_wb.active.cell(row=index, column=17, value=row.wire.wire_type)
        else:
            pass # No wire data is available so leave these cells empty


    # PHASE 3: Export errors to second sheet in workbook

    # FINALLY: Save and wrap up
    output_wb.save(e3_fromto_excel_path)
    
    if verbose:
        click.echo(f"Conversion complete! Output saved to: {e3_fromto_excel_path}")
    else:
        click.echo(f"Successfully converted {len(e3_fromto)} connections to {e3_fromto_excel_path}")


if __name__ == '__main__':
    cli_main()

