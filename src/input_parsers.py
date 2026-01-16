"""Input parsers for different harness CAD formats.

This module provides an abstract base class for input parsing and concrete
implementations for specific formats like RapidHarness. New CAD formats can
be supported by implementing the InputParser interface.
"""

from abc import ABC, abstractmethod
from pathlib import Path
import re
import openpyxl
import click
from colorama import Fore, Style

from models import E3FromToListRow, RapidHarnessEndpoint, ConversionError


class InputParser(ABC):
    """Abstract base class for harness data input parsers."""
    
    @abstractmethod
    def parse(self, input_file: Path, wire_lut: dict, errors: list, verbose: bool = False):
        """Parse input file and return list of E3FromToListRow objects.
        
        Args:
            input_file: Path to input file
            wire_lut: Wire lookup table (maps wire names to E3WireComponent)
            errors: List to accumulate ConversionError objects
            verbose: Enable verbose output
            
        Returns:
            List of E3FromToListRow objects
        """
        pass


class RapidHarnessParser(InputParser):
    """Parser for RapidHarness Excel exports.
    
    Reads from the 'Connections' sheet starting at row 11 with specific
    column mappings.
    """
    
    # Column indices (1-based for cell references, but documented here for clarity)
    COL_FROM_ENDPOINT = 2   # B: FROM Designation (Connector.Pin)
    COL_TO_ENDPOINT = 3     # C: TO Designation (Connector.Pin)
    COL_CONDUCTOR = 4       # D: Conductor (e.g. W19.Black)
    COL_WIRE_SKU = 5        # E: Wire Part Number
    COL_FROM_DEVICE_PN = 11 # K: FROM Connector Part Number
    COL_TO_DEVICE_PN = 13   # M: TO Connector Part Number
    COL_SIGNAL_NAME = 15    # O: Signal Name
    
    HEADER_ROW = 11  # Data starts at row 11
    SHEET_NAME = "Connections"
    
    def parse(self, input_file: Path, wire_lut: dict, errors: list, verbose: bool = False):
        """Parse RapidHarness Excel export.
        
        Args:
            input_file: Path to RapidHarness .xlsx export
            wire_lut: Wire lookup table
            errors: List to accumulate errors
            verbose: Enable verbose output
            
        Returns:
            List of E3FromToListRow objects
        """
        if verbose:
            click.echo("Parsing RapidHarness Excel export...")
        
        e3_fromto = []
        rh_workbook = openpyxl.load_workbook(str(input_file))
        
        if self.SHEET_NAME not in rh_workbook.sheetnames:
            raise click.ClickException(f"Missing '{self.SHEET_NAME}' sheet in RapidHarness export")
        
        sheet = rh_workbook[self.SHEET_NAME]
        
        for row_num in range(self.HEADER_ROW, sheet.max_row + 1):
            e3_fromto_row = E3FromToListRow()
            
            # Column B: FROM Designation (Connector.Pin format)
            from_endpoint = RapidHarnessEndpoint(
                sheet.cell(row=row_num, column=self.COL_FROM_ENDPOINT).value
            )
            e3_fromto_row.from_device_name = from_endpoint.get_device_des()
            e3_fromto_row.from_pin = from_endpoint.get_pin_des()
            
            # Column C: TO Designation (Connector.Pin format)
            to_endpoint = RapidHarnessEndpoint(
                sheet.cell(row=row_num, column=self.COL_TO_ENDPOINT).value
            )
            e3_fromto_row.to_device_name = to_endpoint.get_device_des()
            e3_fromto_row.to_pin = to_endpoint.get_pin_des()
            
            # Column D: Conductor (extract wire index, e.g. "19" from "W19.Black")
            conductor_cell = sheet.cell(row=row_num, column=self.COL_CONDUCTOR).value
            try:
                match = re.search(r'\d+', conductor_cell)
                if match:
                    e3_fromto_row.wire_index = int(match.group())
            except (TypeError, AttributeError):
                pass  # Will be logged if needed in validation
            
            # Column E: Wire Part Number - lookup in wire table
            rh_wire_sku = sheet.cell(row=row_num, column=self.COL_WIRE_SKU).value
            if rh_wire_sku in wire_lut:
                e3_fromto_row.wire = wire_lut[rh_wire_sku]
            elif rh_wire_sku is not None:
                error_msg = f"Wire '{rh_wire_sku}' not found in lookup table"
                click.echo(f"{Fore.RED}[ERROR] Row {row_num}: {error_msg}{Style.RESET_ALL}", err=True)
                error = ConversionError(
                    severity="error",
                    error_type="WIRE_NOT_FOUND",
                    row_number=row_num,
                    entity_id="Wire",
                    entity_value=str(rh_wire_sku),
                    description=error_msg
                )
                errors.append(error)
            
            # Column K: FROM Connector Part Number
            e3_fromto_row.from_device_pn = sheet.cell(row=row_num, column=self.COL_FROM_DEVICE_PN).value
            
            # Column M: TO Connector Part Number
            e3_fromto_row.to_device_pn = sheet.cell(row=row_num, column=self.COL_TO_DEVICE_PN).value
            
            # Column O: Signal Name
            e3_fromto_row.signal_name = sheet.cell(row=row_num, column=self.COL_SIGNAL_NAME).value
            
            e3_fromto.append(e3_fromto_row)
        
        if verbose:
            click.echo(f"Parsed {len(e3_fromto)} connections from RapidHarness")
        
        return e3_fromto
