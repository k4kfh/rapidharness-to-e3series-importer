"""Output writing for E3.series From-To List format."""

from pathlib import Path
import openpyxl
import click
from colorama import Fore, Style

from models import E3FromToListRow, E3WireComponent


# E3.series From-To List column definitions
E3_HEADERS = [
    "From Assignment",
    "From Location",
    "From Device Name",
    "From Device Part #",
    "From Pin",
    "From Pin Part #",
    "To Assignment",
    "To Location",
    "To Device Name",
    "To Device Part #",
    "To Pin",
    "To Pin Part #",
    "Wire/Conductor Number",
    "Signal",
    "Wire Type",
    "Wire Color",
    "Wire Gauge",
]

# Column indices for writing (1-based)
COL_FROM_DEVICE_NAME = 3
COL_FROM_DEVICE_PN = 4
COL_FROM_PIN = 5
COL_TO_DEVICE_NAME = 9
COL_TO_DEVICE_PN = 10
COL_TO_PIN = 11
COL_WIRE_NUMBER = 13
COL_SIGNAL_NAME = 14
COL_WIRE_TYPE = 15
COL_WIRE_COLOR = 16
COL_WIRE_GAUGE = 17


def write_e3_fromto_list(rows: list, output_file: Path):
    """Write E3.series From-To List to Excel file.
    
    Args:
        rows: List of E3FromToListRow objects
        output_file: Path for output .xlsx file
    """
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "From-To List"
    
    # Write headers
    for col_idx, header in enumerate(E3_HEADERS, start=1):
        worksheet.cell(row=1, column=col_idx, value=header)
    
    # Write data rows
    for row_idx, row in enumerate(rows, start=2):
        worksheet.cell(row=row_idx, column=COL_FROM_DEVICE_NAME, value=row.from_device_name)
        worksheet.cell(row=row_idx, column=COL_FROM_DEVICE_PN, value=row.from_device_pn)
        worksheet.cell(row=row_idx, column=COL_FROM_PIN, value=row.from_pin)
        
        worksheet.cell(row=row_idx, column=COL_TO_DEVICE_NAME, value=row.to_device_name)
        worksheet.cell(row=row_idx, column=COL_TO_DEVICE_PN, value=row.to_device_pn)
        worksheet.cell(row=row_idx, column=COL_TO_PIN, value=row.to_pin)
        
        worksheet.cell(row=row_idx, column=COL_WIRE_NUMBER, value=row.wire_index)
        worksheet.cell(row=row_idx, column=COL_SIGNAL_NAME, value=row.signal_name)
        
        # Wire data (only present if not part of cable)
        if isinstance(row.wire, E3WireComponent):
            worksheet.cell(row=row_idx, column=COL_WIRE_TYPE, value=row.wire.wire_group)
            worksheet.cell(row=row_idx, column=COL_WIRE_COLOR, value=row.wire.color)
            worksheet.cell(row=row_idx, column=COL_WIRE_GAUGE, value=row.wire.wire_type)
    
    workbook.save(str(output_file))
